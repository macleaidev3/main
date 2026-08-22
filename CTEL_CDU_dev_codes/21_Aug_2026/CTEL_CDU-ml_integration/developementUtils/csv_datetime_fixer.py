"""
Normalize the "Date"/"Time" column of every CSV and Excel file in a folder.

- Pick a folder (GUI dialog, or pass the path as a command-line argument).
- Every *.csv and *.xlsx/*.xlsm in that folder is read and rewritten into a new
  sub-folder ("modified" by default) using the same file name. Workbooks are
  written out as CSV, so "sheet.xlsx" becomes "modified/sheet.csv"; only the
  first worksheet is read.
- The column named "Date" or "Time" (case-insensitive) is reformatted to
  DD-MM-YYYY HH:MM. Input may use any separator (/ - . space) and may or may
  not have seconds, AM/PM, or a 2-digit year. Date order is assumed to be
  day, month, year -- unless the first component is a 4-digit year (ISO).
- Rows whose date value carries no time part are dropped entirely.
- All other columns are written back exactly as they were read.
"""

import csv
import datetime
import os
import re
import sys

OUTPUT_DIR_NAME = "modified"
TARGET_COLUMNS = {"date", "time"}
OUTPUT_FORMAT = "{day:02d}-{month:02d}-{year:04d} {hour:02d}:{minute:02d}"
EXCEL_EXTENSIONS = (".xlsx", ".xlsm")

# x<sep>y[<sep>z] [ jj:kk[:ll] [am/pm] ]
DATETIME_RE = re.compile(
    r"""^\s*
    (?P<a>\d{1,4}) \s*[/\-.]\s* (?P<b>\d{1,4})
    (?: \s*[/\-.]\s* (?P<c>\d{1,4}) )?
    (?:
        [\sT,]+
        (?P<hour>\d{1,2}) \s*[:.]\s* (?P<minute>\d{1,2})
        (?: \s*[:.]\s* (?P<second>\d{1,2}(?:\.\d+)?) )?
        \s* (?P<ampm>[AaPp]\.?[Mm]\.?)?
    )?
    \s*$""",
    re.VERBOSE,
)


class NoTimeError(ValueError):
    """The value parsed as a date but carried no time component."""


def parse_datetime(raw):
    """Return 'DD-MM-YYYY HH:MM' for a date-time string, day-first."""
    if raw is None:
        raise ValueError("empty value")

    match = DATETIME_RE.match(raw.strip())
    if not match:
        raise ValueError("unrecognized format: %r" % raw)

    if match.group("hour") is None:
        raise NoTimeError(raw)

    a, b, c = match.group("a"), match.group("b"), match.group("c")

    if c is None:
        # Only two date components, e.g. "mm-yyyy HH:MM" -> day is unknown.
        month, year, day = int(a), int(b), 1
    elif len(a) == 4:
        # ISO-ish: yyyy-mm-dd
        year, month, day = int(a), int(b), int(c)
    else:
        day, month, year = int(a), int(b), int(c)

    if year < 100:
        year += 2000 if year < 70 else 1900

    hour = int(match.group("hour"))
    minute = int(match.group("minute"))

    ampm = match.group("ampm")
    if ampm:
        ampm = ampm.lower().replace(".", "")
        if ampm == "pm" and hour < 12:
            hour += 12
        elif ampm == "am" and hour == 12:
            hour = 0

    if not (1 <= month <= 12 and 1 <= day <= 31 and hour <= 23 and minute <= 59):
        raise ValueError("out-of-range date-time: %r" % raw)

    return OUTPUT_FORMAT.format(day=day, month=month, year=year, hour=hour, minute=minute)


def find_target_column(header):
    for index, name in enumerate(header):
        if name.strip().lstrip("﻿").lower() in TARGET_COLUMNS:
            return index
    return None


def sniff_delimiter(path):
    with open(path, "r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(8192)
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        return ","


def read_csv_rows(path):
    """Return (rows, delimiter) for a CSV file."""
    delimiter = sniff_delimiter(path)
    with open(path, "r", encoding="utf-8-sig", newline="") as fin:
        return list(csv.reader(fin, delimiter=delimiter)), delimiter


def cell_to_text(value, number_format=""):
    """Flatten one Excel cell to the text form parse_datetime expects."""
    if value is None:
        return ""
    if isinstance(value, datetime.datetime):
        # A date-only cell still arrives as midnight, so the number format is
        # what separates "no time part" from a real 00:00 timestamp.
        if "h" in number_format.lower() or (value.hour or value.minute or value.second):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        return value.strftime("%Y-%m-%d")
    if isinstance(value, datetime.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, datetime.time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def read_excel_rows(path):
    """Return (rows, delimiter) for the first worksheet of a workbook."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ValueError("openpyxl is not installed; cannot read Excel files")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        rows = [
            [cell_to_text(cell.value, cell.number_format or "") for cell in cells]
            for cells in sheet.iter_rows()
        ]
    finally:
        workbook.close()

    # Trailing empty rows/columns are an Excel artefact, not data.
    while rows and not any(field.strip() for field in rows[0]):
        rows.pop(0)
    if rows:
        header = rows[0]
        while header and not header[-1].strip():
            header.pop()
        width = len(header)
        rows = [row[:width] + [""] * (width - len(row)) for row in rows]

    return rows, ","


def read_rows(path):
    if path.lower().endswith(EXCEL_EXTENSIONS):
        return read_excel_rows(path)
    return read_csv_rows(path)


def process_file(src, dst):
    """Rewrite one CSV or workbook as CSV. Returns (kept, dropped, unparsed)."""
    rows, delimiter = read_rows(src)

    if not rows:
        raise ValueError("file is empty")

    header = rows[0]
    column = find_target_column(header)
    if column is None:
        raise ValueError('no "Date" or "Time" column found')

    kept, dropped, unparsed = [], 0, []

    for line_no, row in enumerate(rows[1:], start=2):
        if not any(field.strip() for field in row):
            continue
        if column >= len(row):
            dropped += 1
            continue
        try:
            row[column] = parse_datetime(row[column])
        except NoTimeError:
            dropped += 1
            continue
        except ValueError as err:
            unparsed.append((line_no, str(err)))
            dropped += 1
            continue
        kept.append(row)

    with open(dst, "w", encoding="utf-8-sig", newline="") as fout:
        writer = csv.writer(fout, delimiter=delimiter, lineterminator="\n")
        writer.writerow(header)
        writer.writerows(kept)

    return len(kept), dropped, unparsed


def choose_folder():
    if len(sys.argv) > 1:
        return sys.argv[1]
    try:
        import tkinter as tk
        from tkinter import filedialog
    except ImportError:
        return input("Folder path: ").strip('" ')

    root = tk.Tk()
    root.withdraw()
    folder = filedialog.askdirectory(title="Select the folder containing the CSV/Excel files")
    root.destroy()
    return folder


def main():
    folder = choose_folder()
    if not folder or not os.path.isdir(folder):
        print("No valid folder selected.")
        return 1

    out_dir = os.path.join(folder, OUTPUT_DIR_NAME)
    os.makedirs(out_dir, exist_ok=True)

    sources = sorted(
        name for name in os.listdir(folder)
        if name.lower().endswith((".csv",) + EXCEL_EXTENSIONS)
        and not name.startswith("~$")  # Excel lock file
        and os.path.isfile(os.path.join(folder, name))
    )
    if not sources:
        print("No CSV or Excel files found in %s" % folder)
        return 1

    print("Reading from : %s" % folder)
    print("Writing to   : %s\n" % out_dir)

    written = {}

    for name in sources:
        src = os.path.join(folder, name)
        out_name = os.path.splitext(name)[0] + ".csv"
        dst = os.path.join(out_dir, out_name)

        if out_name in written:
            print("%-40s SKIPPED (would overwrite the output of %s)"
                  % (name, written[out_name]))
            continue

        try:
            kept, dropped, unparsed = process_file(src, dst)
        except (ValueError, OSError, UnicodeDecodeError) as err:
            print("%-40s SKIPPED (%s)" % (name, err))
            continue

        written[out_name] = name

        print("%-40s %d rows kept, %d dropped" % (name, kept, dropped))
        for line_no, reason in unparsed[:5]:
            print("    line %d: %s" % (line_no, reason))
        if len(unparsed) > 5:
            print("    ... and %d more unparsed values" % (len(unparsed) - 5))

    print("\nDone.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
