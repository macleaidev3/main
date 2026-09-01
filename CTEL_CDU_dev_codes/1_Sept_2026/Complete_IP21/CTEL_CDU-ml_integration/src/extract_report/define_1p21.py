import io
import copy
import traceback
import logging
from pathlib import Path
from src.utils.table_columns import TABLE_COLUMNS
from src.server_manager.operation_manager import DatabaseManager
from src.utils.core_utility_functions import extract_column_names, month_short_name
from datetime import datetime
from openpyxl import load_workbook

# Retrieve the dedicated application logger
logger = logging.getLogger("SentinelApp")

class IP21Exporter:
    def __init__(self, selected_ip21, start_year, end_year, start_month, end_month, temp_dir, template_bytes):
        self.selected_ip21 = selected_ip21
        self.start_year = start_year
        self.end_year = end_year
        self.start_month = start_month
        self.end_month = end_month
        self.temp_dir = temp_dir
        self.month_short_names = month_short_name()
        self.template_bytes = template_bytes

    def _build_year_workbook_file(self, year, column_names, first_month, last_month, temp_dir=None):
        """
        Create one XLSX workbook for a single year using the template.
        Each month gets its own sheet with logos and formatting preserved.
        """
        db_manager = DatabaseManager()
        db_name = "SentinelDB"
        
        wb = None
        out_path = Path(temp_dir) / f"{year}.xlsx" if temp_dir else None

        logger.debug("Starting build for IP21 workbook year: %s", year)

        try:
            # 1. Load the workbook from the in-memory byte stream
            try:
                wb = load_workbook(io.BytesIO(self.template_bytes))
            except Exception as e:
                logger.exception("Failed to load IP21 template from memory.")
                raise RuntimeError(f"Failed to load IP21 template from memory: {e}")
            
            template_ws = wb.active 
            generation_date = datetime.now().strftime("%Y-%m-%d")
            
            # 2. Iterate through the months
            for month in range(first_month, last_month + 1):
                month_name = self.month_short_names[month - 1]
                table_name = f"ip21_{year}_{month_name}"
                
                safe_sheet_name = month_name[:31]
                logger.debug("Generating sheet '%s' for table '%s'.", safe_sheet_name, table_name)

                # Duplicate the template sheet
                ws = wb.copy_worksheet(template_ws)
                ws.title = safe_sheet_name

                # Copy logos/images securely
                if hasattr(template_ws, '_images') and template_ws._images:
                    for img in template_ws._images:
                        ws.add_image(copy.deepcopy(img))

                # 3. Populate Metadata
                ws["B2"] = generation_date
                ws["B3"] = "IP21 Data" 

                # 4. Populate Headers (Row 5)
                for col_idx, header_name in enumerate(column_names, start=1):
                    ws.cell(row=5, column=col_idx, value=header_name)

                # 5. Populate Data starting from Row 6
                current_row = 6
                try:
                    for chunk in db_manager.read_table_by_chunks(str(db_name), table_name, chunk_size=1000):
                        for row_data in chunk:
                            for col_idx, cell_value in enumerate(row_data, start=1):
                                ws.cell(row=current_row, column=col_idx, value=cell_value)
                            current_row += 1
                            
                except Exception as chunk_exc:
                    # Gracefully handle months where the table hasn't been created yet
                    error_str = str(chunk_exc)
                    if '42S02' in error_str or 'Invalid object name' in error_str:
                        logger.info("No IP21 data recorded in %s %s. Table skipped.", month_name, year)
                    else:
                        logger.warning("Unexpected database error while reading %s: %s", table_name, chunk_exc)

            # 6. Cleanup: Remove the original template sheet
            wb.remove(template_ws)

            # 7. Save to disk
            if not out_path:
                raise ValueError("temp_dir was not provided; cannot save the workbook.")
            
            out_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(out_path)
            
            logger.info("Successfully saved IP21 workbook for year %s at %s", year, out_path)
            return out_path

        # --- Top-Level Error Handling ---
        except PermissionError as pe:
            logger.error("Permission denied when saving %s: %s", out_path, pe)
            raise
        except OSError as ose:
            logger.error("OS error occurred while generating IP21 report (%s): %s", year, ose)
            raise
        except Exception as e:
            logger.exception("Critical error generating workbook for IP21 (%s): %s", year, e)
            raise RuntimeError(f"Report generation failed for IP21 ({year})") from e

        # --- Absolute Memory Guarantee ---
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception as close_exc:
                    logger.error("Failed to close IP21 workbook cleanly (%s): %s", year, close_exc)


    def generate_files(self):
        logger.info("Starting batch IP21 file generation.")
        # if self.selected_ip21 is not empty, we will generate files for it. Otherwise, we skip this section.
        if not self.selected_ip21:
            logger.debug("No IP21 selections found. Skipping generation.")
            return
        
        column_names = extract_column_names(TABLE_COLUMNS["ip21_data"])
    
        for year in range(self.start_year, self.end_year + 1):
            first_month = self.start_month if year == self.start_year else 1
            last_month = self.end_month if year == self.end_year else 12

            xlsx_path = self._build_year_workbook_file(
                year=year,
                column_names=column_names,
                first_month=first_month,
                last_month=last_month,
                temp_dir=self.temp_dir,
            )
            archive_name = f"IP21/{year}.xlsx"
            yield archive_name, xlsx_path