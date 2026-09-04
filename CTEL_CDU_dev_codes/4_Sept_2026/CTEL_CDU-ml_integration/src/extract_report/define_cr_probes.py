import traceback
import logging
from datetime import datetime
import copy
import io
from openpyxl import load_workbook
from pathlib import Path
from src.utils.table_columns import TABLE_COLUMNS
from src.server_manager.operation_manager import DatabaseManager
from src.utils.core_utility_functions import extract_column_names, month_short_name

# Retrieve the dedicated application logger
logger = logging.getLogger("SentinelApp")

class CorrosionProbeExporter:
    def __init__(self, selected_probes, start_year, end_year, start_month, end_month, temp_dir, template_bytes):
        self.selected_probes = selected_probes
        self.start_year = start_year
        self.end_year = end_year
        self.start_month = start_month
        self.end_month = end_month
        self.temp_dir = temp_dir
        self.month_short_names = month_short_name()
        self.template_bytes = template_bytes

    def _build_year_workbook_file(self, probe_id, year, column_names, first_month, last_month, temp_dir=None):
        """
        Create one XLSX workbook for a single probe and year using a template.
        Wrapped in comprehensive error handling to guarantee memory cleanup.
        """
        db_manager = DatabaseManager()
        db_name = "SentinelDB"
        
        # Initialize variables before the try block so the 'finally' block can access them safely
        wb = None
        out_path = Path(temp_dir) / f"ID_{probe_id}_{year}.xlsx" if temp_dir else None

        logger.info("Generating workbook for Probe ID: %s | Year: %s", probe_id, year)

        try:
            # 1. Load the workbook from the in-memory byte stream
            try:
                wb = load_workbook(io.BytesIO(self.template_bytes))
            except Exception as e:
                logger.exception("Failed to load template from memory.")
                raise RuntimeError(f"Failed to load template from memory: {e}")
            
            template_ws = wb.active 
            generation_date = datetime.now().strftime("%Y-%m-%d")
            
            # 2. Iterate through the months
            for month in range(first_month, last_month + 1):
                month_name = self.month_short_names[month - 1]
                table_name = f"ut_{probe_id}_{year}_{month_name}_thickness"
                
                safe_sheet_name = month_name[:31]

                # Duplicate the template sheet
                ws = wb.copy_worksheet(template_ws)
                ws.title = safe_sheet_name

                # Copy logos/images
                if hasattr(template_ws, '_images') and template_ws._images:
                    for img in template_ws._images:
                        ws.add_image(copy.deepcopy(img))

                # 3. Populate Metadata
                ws["B2"] = generation_date
                ws["B3"] = f"Probe {probe_id}" 

                # 4. Populate Headers
                for col_idx, header_name in enumerate(column_names, start=1):
                    ws.cell(row=5, column=col_idx, value=header_name)

                # 5. Populate Data
                current_row = 6
                try:
                    for chunk in db_manager.read_table_by_chunks(str(db_name), table_name, chunk_size=1000):
                        for row_data in chunk:
                            for col_idx, cell_value in enumerate(row_data, start=1):
                                ws.cell(row=current_row, column=col_idx, value=cell_value)
                            current_row += 1
                            
                except Exception as chunk_exc:
                    # Convert the exception to a string to safely inspect the ODBC error codes
                    error_str = str(chunk_exc)
                    
                    # '42S02' is the standard SQL state for "Base table or view not found"
                    if '42S02' in error_str or 'Invalid object name' in error_str:
                        # This is a normal business case: No data for this month.
                        logger.info("No data recorded for Probe %s in %s %s. Table skipped.", probe_id, month_name, year)
                    else:
                        # This is a genuine, unexpected database error (e.g., connection lost)
                        logger.warning("Unexpected database error while reading %s: %s", table_name, chunk_exc)

            # 6. Cleanup: Remove the original template sheet
            wb.remove(template_ws)

            # 7. Save to disk
            if not out_path:
                raise ValueError("temp_dir was not provided; cannot save the workbook.")
            
            out_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(out_path)
            
            logger.info("Workbook saved successfully: %s", out_path)
            return out_path

        # --- Top-Level Error Handling ---
        except PermissionError as pe:
            logger.error("Permission denied when saving %s: %s", out_path, pe)
            raise
        except OSError as ose:
            logger.error("OS error occurred while generating report for %s (%s): %s", probe_id, year, ose)
            raise
        except Exception as e:
            # Catch-all for unexpected crashes
            logger.exception("Critical error generating workbook for probe %s (%s)", probe_id, year)
            raise RuntimeError(f"Report generation failed for {probe_id} ({year})") from e

        # --- Absolute Memory Guarantee ---
        finally:
            if wb is not None:
                try:
                    # wb.close() unbinds internal caches in openpyxl, freeing RAM immediately
                    wb.close()
                except Exception as close_exc:
                    logger.error("Failed to close workbook cleanly for probe %s (%s): %s", probe_id, year, close_exc)

    def generate_files(self):
        logger.info("Starting batch export for corrosion probes.")
        if not self.selected_probes:
            return
        
        column_names = extract_column_names(TABLE_COLUMNS["ut_thickness"])

        for probe_id in self.selected_probes:
            for year in range(self.start_year, self.end_year + 1):
                first_month = self.start_month if year == self.start_year else 1
                last_month = self.end_month if year == self.end_year else 12

                xlsx_path = self._build_year_workbook_file(
                    probe_id=probe_id,
                    year=year,
                    column_names=column_names,
                    first_month=first_month,
                    last_month=last_month,
                    temp_dir=self.temp_dir,
                )
                archive_name = f"corrosion probes/{probe_id}/ID_{probe_id}_{year}.xlsx"
                yield archive_name, xlsx_path