import io
import copy
import traceback
import logging
from datetime import datetime
from openpyxl import load_workbook
from pathlib import Path
from src.utils.table_columns import TABLE_COLUMNS
from src.server_manager.operation_manager import DatabaseManager
from src.utils.core_utility_functions import extract_column_names

# Retrieve the dedicated application logger
logger = logging.getLogger("SentinelApp")

class CrudeExporter:
    def __init__(self, selected_crude, start_year, end_year, start_month, end_month, temp_dir, template_bytes):
        self.selected_crude = selected_crude
        self.temp_dir = temp_dir
        self.template_bytes = template_bytes

    def _build_year_workbook_file(self, column_names, temp_dir=None):
        """
        Create one XLSX workbook for the general crude data using the template.
        Uses a single sheet but retains all logos and formatting.
        """
        db_manager = DatabaseManager()
        db_name = "SentinelDB"
        table_name = "crude_data"
        
        wb = None
        out_path = Path(temp_dir) / "General Crude.xlsx" if temp_dir else None

        logger.info("Generating workbook for General Crude data.")

        try:
            # 1. Load the workbook from the in-memory byte stream
            try:
                wb = load_workbook(io.BytesIO(self.template_bytes))
            except Exception as e:
                logger.exception("Failed to load General Crude template from memory.")
                raise RuntimeError(f"Failed to load General Crude template from memory: {e}")
            
            template_ws = wb.active 
            generation_date = datetime.now().strftime("%Y-%m-%d")
            
            # Duplicate the template sheet for the single Crude sheet
            ws = wb.copy_worksheet(template_ws)
            ws.title = "Crude"[:31]

            # Copy logos/images securely
            if hasattr(template_ws, '_images') and template_ws._images:
                for img in template_ws._images:
                    ws.add_image(copy.deepcopy(img))

            # 2. Populate Metadata
            ws["B2"] = generation_date
            ws["B3"] = "General Crude Data" 

            # 3. Populate Headers (Row 5)
            for col_idx, header_name in enumerate(column_names, start=1):
                ws.cell(row=5, column=col_idx, value=header_name)

            # 4. Populate Data starting from Row 6
            current_row = 6
            try:
                for chunk in db_manager.read_table_by_chunks(str(db_name), table_name, chunk_size=1000):
                    for row_data in chunk:
                        for col_idx, cell_value in enumerate(row_data, start=1):
                            ws.cell(row=current_row, column=col_idx, value=cell_value)
                        current_row += 1
                        
            except Exception as chunk_exc:
                # Gracefully handle if the crude_data table hasn't been created yet
                error_str = str(chunk_exc)
                if '42S02' in error_str or 'Invalid object name' in error_str:
                    logger.info("No General Crude data recorded (table '%s' not found).", table_name)
                else:
                    logger.warning("Unexpected database error while reading %s: %s", table_name, chunk_exc)

            # 5. Cleanup: Remove the original blank template sheet
            wb.remove(template_ws)

            # 6. Save to disk
            if not out_path:
                raise ValueError("temp_dir was not provided; cannot save the workbook.")
            
            out_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(out_path)
            
            logger.info("General Crude workbook saved successfully at: %s", out_path)
            return out_path

        # --- Top-Level Error Handling ---
        except PermissionError as pe:
            logger.error("Permission denied when saving %s: %s", out_path, pe)
            raise
        except OSError as ose:
            logger.error("OS error occurred while generating General Crude report: %s", ose)
            raise
        except Exception as e:
            logger.exception("Critical error generating workbook for General Crude.")
            raise RuntimeError("Report generation failed for General Crude") from e

        # --- Absolute Memory Guarantee ---
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception as close_exc:
                    logger.error("Failed to close General Crude workbook cleanly: %s", close_exc)


    def generate_files(self):
        logger.info("Starting General Crude export.")
        # if self.selected_crude is not empty, we will generate files for it. Otherwise, we skip this section.
        if not self.selected_crude:
            logger.debug("No crude selection found. Skipping generation.")
            return
        
        column_names = extract_column_names(TABLE_COLUMNS["crude_data"])
    
        xlsx_path = self._build_year_workbook_file(
                column_names=column_names,
                temp_dir=self.temp_dir,
            )
        
        archive_name = "Crude/crude.xlsx"
        yield archive_name, xlsx_path