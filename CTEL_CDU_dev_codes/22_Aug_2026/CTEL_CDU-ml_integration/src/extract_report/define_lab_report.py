import io
import copy
import traceback
import logging
from datetime import datetime
from openpyxl import load_workbook
from pathlib import Path
from src.utils.table_columns import TABLE_COLUMNS
from src.server_manager.operation_manager import DatabaseManager
from src.utils.core_utility_functions import extract_column_names, month_short_name

# Retrieve the dedicated application logger
logger = logging.getLogger("SentinelApp")

class LabReportExporter:
    def __init__(self, selected_lab_results, start_year, end_year, start_month, end_month, temp_dir, template_bytes):
        
        self.selected_lab_results = selected_lab_results
        self.start_year = start_year
        self.end_year = end_year
        self.start_month = start_month
        self.end_month = end_month
        self.temp_dir = temp_dir
        self.month_short_names = month_short_name()
        self.template_bytes = template_bytes

    def _build_year_workbook_file(self, lab_test, display_name, year, column_names, first_month, last_month, temp_dir=None):
        """
        Create one XLSX workbook for a single lab test and year using the template.
        Each month gets its own sheet with logos and formatting preserved.
        """
        db_manager = DatabaseManager()
        db_name = "SentinelDB" 
        
        wb = None
        out_path = Path(temp_dir) / f"{year}.xlsx" if temp_dir else None

        logger.info("Generating Lab Report workbook for: %s | Year: %s", display_name, year)

        try:
            # 1. Load the workbook from the in-memory byte stream
            try:
                wb = load_workbook(io.BytesIO(self.template_bytes))
            except Exception as e:
                logger.exception("Failed to load Lab Report template from memory.")
                raise RuntimeError(f"Failed to load Lab Report template from memory: {e}")
            
            template_ws = wb.active 
            generation_date = datetime.now().strftime("%Y-%m-%d")
            
            # 2. Iterate through the months
            for month in range(first_month, last_month + 1):
                month_name = self.month_short_names[month - 1]
                table_name = f"lab_{year}_{month_name}_{lab_test}"
                
                safe_sheet_name = month_name[:31]
                logger.debug("Generating sheet '%s' for table '%s'.", safe_sheet_name, table_name)

                # Duplicate the template sheet
                ws = wb.copy_worksheet(template_ws)
                ws.title = safe_sheet_name

                # Copy logos/images securely
                if hasattr(template_ws, '_images') and template_ws._images:
                    for img in template_ws._images:
                        ws.add_image(copy.deepcopy(img))

                # 3. Populate Metadata
                ws["B2"] = generation_date
                ws["B3"] = f"Lab Report: {display_name}" # Uses the clean UI string

                # 4. Populate Headers (Row 5)
                for col_idx, header_name in enumerate(column_names, start=1):
                    ws.cell(row=5, column=col_idx, value=header_name)

                # 5. Populate Data starting from Row 6
                current_row = 6
                try:
                    for chunk in db_manager.read_table_by_chunks(str(db_name), table_name, chunk_size=1000):
                        for row_data in chunk:
                            for col_idx, cell_value in enumerate(row_data, start=1):
                                ws.cell(row=current_row, column=col_idx, value=cell_value)
                            current_row += 1
                            
                except Exception as chunk_exc:
                    # Gracefully handle months where the table hasn't been created yet
                    error_str = str(chunk_exc)
                    if '42S02' in error_str or 'Invalid object name' in error_str:
                        logger.info("No data recorded for '%s' in %s %s. Table skipped.", display_name, month_name, year)
                    else:
                        logger.warning("Unexpected database error while reading %s: %s", table_name, chunk_exc)

            # 6. Cleanup: Remove the original template sheet
            wb.remove(template_ws)

            # 7. Save to disk
            if not out_path:
                raise ValueError("temp_dir was not provided; cannot save the workbook.")
            
            out_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(out_path)
            
            logger.info("Lab report saved successfully: %s", out_path)
            return out_path

        # --- Top-Level Error Handling ---
        except PermissionError as pe:
            logger.error("Permission denied when saving %s: %s", out_path, pe)
            raise
        except OSError as ose:
            logger.error("OS error occurred while generating Lab report for %s (%s): %s", display_name, year, ose)
            raise
        except Exception as e:
            logger.exception("Critical error generating workbook for %s (%s)", display_name, year)
            raise RuntimeError(f"Report generation failed for {display_name} ({year})") from e

        # --- Absolute Memory Guarantee ---
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception as close_exc:
                    logger.error("Failed to close Lab workbook cleanly for %s (%s): %s", display_name, year, close_exc)


    def generate_files(self):
        logger.info("Starting batch Lab Report export.")
        if not self.selected_lab_results:
            logger.debug("No lab result selections found. Skipping generation.")
            return
        
        logger.debug("Selected lab results: %s", self.selected_lab_results)
        
        for original_lab_test in self.selected_lab_results:
            # We save the original string to pass as the display_name
            display_name = original_lab_test 
            
            if original_lab_test == 'After Desalter Stage 1':
                column_names = extract_column_names(TABLE_COLUMNS["after_desalter_stage_1"])
                lab_test = "after_desalter_stage_1"
            elif original_lab_test == 'After Desalter Stage 2':
                column_names = extract_column_names(TABLE_COLUMNS["after_desalter_stage_2"])
                lab_test = "after_desalter_stage_2"
            elif original_lab_test == 'Crude Before Desalter':
                column_names = extract_column_names(TABLE_COLUMNS["crude_before_desalter"])
                lab_test = "crude_before_desalter"
            elif original_lab_test == 'Sour Water ICV 112':
                column_names = extract_column_names(TABLE_COLUMNS["sour_water_icv112"])
                lab_test = "sour_water_icv112"
            elif original_lab_test == 'Sour Water ICV 113':
                column_names = extract_column_names(TABLE_COLUMNS["sour_water_icv113"])
                lab_test = "sour_water_icv113"
            elif original_lab_test == "Stripped Water":
                column_names = extract_column_names(TABLE_COLUMNS["stripped_water"])
                lab_test = "stripped_water"
            else:
                # Fallback if a new test gets added to the UI but not this loop
                logger.warning("Unknown lab test encountered: %s. Skipping.", original_lab_test)
                continue 

            for year in range(self.start_year, self.end_year + 1):
                first_month = self.start_month if year == self.start_year else 1
                last_month = self.end_month if year == self.end_year else 12

                xlsx_path = self._build_year_workbook_file(
                    lab_test=lab_test,
                    display_name=display_name, # Passed down for cell B3
                    year=year,
                    column_names=column_names,
                    first_month=first_month,
                    last_month=last_month,
                    temp_dir=self.temp_dir,
                )
                archive_name = f"Lab reports/{lab_test}/{year}.xlsx"
                yield archive_name, xlsx_path